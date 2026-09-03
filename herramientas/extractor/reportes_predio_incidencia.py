from __future__ import annotations

import re
import unicodedata
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter


KEY_COLS = ["Numero_Predio", "Numero_Incidencia"]


def clean(value) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ").strip())


def normalized(value) -> str:
    text = clean(value).upper()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_predio(value) -> str:
    match = re.search(r"\d{6,8}", clean(value))
    return match.group(0) if match else ""


def clean_incidencia(value) -> str:
    match = re.search(r"NI-\d+", clean(value), re.IGNORECASE)
    return match.group(0).upper() if match else ""


def first_nonempty(df: pd.DataFrame, columns: list[str]) -> pd.Series:
    out = pd.Series([""] * len(df), index=df.index, dtype=object)
    for col in columns:
        if col not in df.columns:
            continue
        values = df[col].map(clean)
        out = out.mask(out.eq("") & values.ne(""), values)
    return out


def load_assignment_source(path: str | Path | None) -> dict[tuple[str, str], dict[str, str]]:
    if not path:
        return {}
    path = Path(path)
    if not path.exists():
        return {}

    assignments: dict[tuple[str, str], dict[str, str]] = {}
    try:
        xl = pd.ExcelFile(path)
    except Exception:
        return assignments

    sheets = [s for s in xl.sheet_names if s.startswith("Asignado -")]
    sheets += [s for s in xl.sheet_names if s not in sheets and s.lower().startswith("predios")]
    for sheet_name in sheets:
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, dtype=str, keep_default_na=False).fillna("")
        except Exception:
            continue
        if "Predio" not in df.columns or "Incidencia" not in df.columns:
            continue

        sheet_assignment = ""
        if sheet_name.startswith("Asignado -"):
            sheet_assignment = clean(sheet_name.replace("Asignado -", "")) or "Sin asignar"

        for _, row in df.iterrows():
            predio = clean_predio(row.get("Predio", ""))
            incidencia = clean_incidencia(row.get("Incidencia", ""))
            if not predio or not incidencia:
                continue
            assigned = clean(row.get("Asignados", "")) or sheet_assignment
            if normalized(assigned) == "SIN ASIGNAR":
                assigned = "Sin asignar"
            if not assigned:
                continue
            assignments.setdefault(
                (predio, incidencia),
                {
                    "Asignado_Original": assigned,
                    "Hoja_Asignado_Original": sheet_name,
                },
            )
    return assignments


def enrich_main_df(main_df: pd.DataFrame, assignment_source: str | Path | None = None) -> pd.DataFrame:
    df = main_df.copy().fillna("")
    assignments = load_assignment_source(assignment_source)

    predio = first_nonempty(df, ["Numero_Predio"])
    incidencia = first_nonempty(df, ["Numero_Incidencia"])
    fallback_assigned = first_nonempty(
        df,
        [
            "Asignado_Original",
            "Origen_Asignados",
            "Origen_Asignado",
            "Origen_Tecnico",
            "Origen_Tecnicos",
            "Origen_Responsable",
            "Origen_Instalador",
        ],
    )

    assigned = []
    assigned_sheet = []
    for predio_value, incidencia_value, fallback_value in zip(predio, incidencia, fallback_assigned):
        item = assignments.get((clean_predio(predio_value), clean_incidencia(incidencia_value)), {})
        value = clean(item.get("Asignado_Original", "")) or fallback_value
        if normalized(value) == "SIN ASIGNAR":
            value = "Sin asignar"
        assigned.append(value or "Sin dato")
        assigned_sheet.append(item.get("Hoja_Asignado_Original", ""))

    # El default tiene que ser una Serie, no un 0: si la columna no existe -pasa cuando
    # ningun predio del lote tiene cronogramas- df.get devolvia un int y el .fillna de
    # abajo reventaba. Las otras dos llamadas a df.get del archivo ya lo hacian asi.
    cron_count = pd.to_numeric(
        df.get("Cronogramas_Originados_Cantidad", pd.Series(0, index=df.index)),
        errors="coerce",
    ).fillna(0).astype(int)
    cron_name = first_nonempty(df, ["Cronograma_Ultimo_Nombre_de_Cronograma", "Incidencia_Cronograma"])
    estado = first_nonempty(df, ["Incidencia_Estado"])

    df["Asignado_Original"] = assigned
    df["Hoja_Asignado_Original"] = assigned_sheet
    df["Zona_Reporte"] = first_nonempty(df, ["Predio_Zona_LAC_correspondiente", "Origen_Zona", "Origen_Zona_LAC"])
    df["Provincia_Reporte"] = first_nonempty(df, ["Predio_Plan_Provincia", "Origen_Provincia", "Incidencia_Provincia"])
    df["Departamento_Reporte"] = first_nonempty(df, ["Predio_Departamento", "Origen_Departamento", "Incidencia_Departamento"])
    df["Codigo_Localidad_Reporte"] = first_nonempty(df, ["Predio_Codigo_Localidad", "Origen_Codigo_Localidad"])
    df["Direccion_Reporte"] = first_nonempty(df, ["Predio_Direccion", "Origen_Direccion", "Incidencia_Direccion"])
    df["GPS_Reporte"] = first_nonempty(
        df,
        [
            "Predio_Coordenadas_GPS",
            "Predio_Coordenadas_GPS_Instalador",
            "Predio_Coordenadas_GPS_Territorio",
            "Origen_GPS",
            "Incidencia_Coordenadas_GPS_Instalador",
        ],
    )
    df["Institucion_Reporte"] = first_nonempty(
        df,
        ["Origen_Institucion", "Incidencia_Nombre_Escuela", "Predio_Nombre_de_la_cuenta"],
    )
    df["CUE_Reporte"] = first_nonempty(
        df,
        ["Predio_CUE_Predio", "Origen_CUE_Predio", "Origen_CUE", "Predio_Consolidado_de_CUEs"],
    )
    df["Estado_Incidencia_Reporte"] = estado
    df["Estado_Nivel3_Reporte"] = first_nonempty(df, ["Incidencia_Estado_de_Nivel_3"])
    df["Motivo_Nivel3_Reporte"] = first_nonempty(
        df,
        [
            "Incidencia_Motivo_en_tratamiento_Estado_Nivel_3",
            "Incidencia_Motivo_resuelto_Estado_Nivel_3",
            "Origen_Motivo_resuelto_Estado_Nivel_3",
        ],
    )
    df["Comentario_Nivel3_Reporte"] = first_nonempty(df, ["Incidencia_Comentario_Nivel_3"])
    df["Tiene_Cronograma"] = (cron_count.gt(0) | cron_name.ne("")).map(lambda value: "SI" if value else "NO")
    df["Cronograma_Reporte"] = cron_name
    df["Cronograma_Inicio_Reporte"] = first_nonempty(df, ["Cronograma_Ultimo_Fecha_de_Inicio", "Origen_DESDE"])
    df["Cronograma_Fin_Reporte"] = first_nonempty(df, ["Cronograma_Ultimo_Fecha_de_Fin", "Origen_HASTA"])
    df["Ultimo_Cronograma_Estado"] = first_nonempty(df, ["Cronograma_Ultimo_Estado"])

    fecha_referencia = pd.Timestamp(date.today())
    inicio_cronograma = pd.to_datetime(df["Cronograma_Inicio_Reporte"], dayfirst=True, errors="coerce")
    fin_cronograma = pd.to_datetime(df["Cronograma_Fin_Reporte"], dayfirst=True, errors="coerce")
    estado_cronograma = df["Ultimo_Cronograma_Estado"].map(normalized)
    situacion_cronograma = []
    for tiene_cronograma, estado_crono, inicio, fin in zip(
        df["Tiene_Cronograma"], estado_cronograma, inicio_cronograma, fin_cronograma
    ):
        if tiene_cronograma != "SI":
            situacion_cronograma.append("SIN CRONOGRAMA")
        elif estado_crono != "APROBADO":
            situacion_cronograma.append(
                "ESTADO NO ACTIVO" if not estado_crono else f"ESTADO {estado_crono}"
            )
        elif pd.isna(inicio) or pd.isna(fin):
            situacion_cronograma.append("SIN FECHAS COMPLETAS")
        elif inicio <= fecha_referencia <= fin:
            situacion_cronograma.append("ACTIVO")
        elif fecha_referencia < inicio:
            situacion_cronograma.append("FUTURO")
        else:
            situacion_cronograma.append("FINALIZADO")
    df["Fecha_Referencia_Cronograma"] = fecha_referencia.strftime("%d/%m/%Y")
    df["Ultimo_Cronograma_Situacion"] = situacion_cronograma
    df["Ultimo_Cronograma_Activo"] = df["Ultimo_Cronograma_Situacion"].eq("ACTIVO").map(
        lambda value: "SI" if value else "NO"
    )
    df["Predio_Verificado"] = first_nonempty(df, ["Salesforce_Id"]).map(lambda value: "SI" if clean(value) else "NO")
    df["Incidencia_Verificada"] = first_nonempty(df, ["Incidencia_Id"]).map(lambda value: "SI" if clean(value) else "NO")
    df["Incidencia_Cerrada"] = estado.map(lambda value: "SI" if normalized(value) == "CERRADO" else "NO")
    return df


def report_columns_first(df: pd.DataFrame) -> pd.DataFrame:
    first_cols = [
        "Numero_Predio",
        "Numero_Incidencia",
        "Asignado_Original",
        "Hoja_Asignado_Original",
        "Zona_Reporte",
        "Provincia_Reporte",
        "Departamento_Reporte",
        "Codigo_Localidad_Reporte",
        "Direccion_Reporte",
        "GPS_Reporte",
        "Institucion_Reporte",
        "CUE_Reporte",
        "Estado_Incidencia_Reporte",
        "Estado_Nivel3_Reporte",
        "Motivo_Nivel3_Reporte",
        "Comentario_Nivel3_Reporte",
        "Tiene_Cronograma",
        "Cronograma_Reporte",
        "Cronograma_Inicio_Reporte",
        "Cronograma_Fin_Reporte",
        "Ultimo_Cronograma_Estado",
        "Fecha_Referencia_Cronograma",
        "Ultimo_Cronograma_Situacion",
        "Ultimo_Cronograma_Activo",
        "Incidencia_Cerrada",
        "Predio_Verificado",
        "Incidencia_Verificada",
        "Salesforce_Id",
        "Incidencia_Id",
        "Predio_Error",
        "Incidencia_Error",
    ]
    cols = [c for c in first_cols if c in df.columns]
    return df[cols + [c for c in df.columns if c not in cols]]


def enrich_related_dfs(related_dfs: dict[str, pd.DataFrame], main_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    context_cols = [
        "Asignado_Original",
        "Zona_Reporte",
        "Provincia_Reporte",
        "Departamento_Reporte",
        "Estado_Incidencia_Reporte",
        "Estado_Nivel3_Reporte",
        "Tiene_Cronograma",
        "Cronograma_Inicio_Reporte",
        "Cronograma_Fin_Reporte",
        "Institucion_Reporte",
    ]
    context = main_df[KEY_COLS + [c for c in context_cols if c in main_df.columns]].drop_duplicates(KEY_COLS)
    out: dict[str, pd.DataFrame] = {}
    for name, rel_df in related_dfs.items():
        rel = rel_df.copy().fillna("")
        if rel.empty or not all(c in rel.columns for c in KEY_COLS):
            out[name] = rel
            continue
        missing_context = [c for c in context.columns if c not in KEY_COLS and c not in rel.columns]
        if missing_context:
            rel = rel.merge(context[KEY_COLS + missing_context], on=KEY_COLS, how="left")
        first_cols = KEY_COLS + [c for c in missing_context if c in rel.columns] + [
            c for c in ["Salesforce_Id", "Incidencia_Id", "Orden_Relacionado"] if c in rel.columns
        ]
        out[name] = rel[[c for c in first_cols if c in rel.columns] + [c for c in rel.columns if c not in first_cols]]
    return out


def _summary_by(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=cols + ["Total", "Incidencia_Verificada", "No_Cerradas", "Cerradas", "Con_Cronograma", "Sin_Cronograma", "Sin_Incidencia_Id"])
    verified = df.get("Incidencia_Verificada", pd.Series("", index=df.index)).eq("SI")
    closed = verified & df["Incidencia_Cerrada"].eq("SI")
    has_cron = df["Tiene_Cronograma"].eq("SI")
    open_known = verified & (~closed)
    tmp = df.assign(
        Total=1,
        Incidencia_Verificada=verified.astype(int),
        No_Cerradas=open_known.astype(int),
        Cerradas=closed.astype(int),
        Con_Cronograma=(open_known & has_cron).astype(int),
        Sin_Cronograma=(open_known & (~has_cron)).astype(int),
        Sin_Incidencia_Id=(~verified).astype(int),
    )
    return (
        tmp.groupby(cols, dropna=False, as_index=False)[
            ["Total", "Incidencia_Verificada", "No_Cerradas", "Cerradas", "Con_Cronograma", "Sin_Cronograma", "Sin_Incidencia_Id"]
        ]
        .sum()
        .sort_values(cols, kind="stable")
    )


def build_extra_sheets(main_df: pd.DataFrame, related_dfs: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    required = {"Tiene_Cronograma", "Predio_Verificado", "Incidencia_Verificada"}
    df = enrich_main_df(main_df) if not required.issubset(main_df.columns) else main_df.copy().fillna("")
    verified = df.get("Incidencia_Verificada", pd.Series("", index=df.index)).eq("SI")
    closed = verified & df["Incidencia_Cerrada"].eq("SI")
    has_cron = df["Tiene_Cronograma"].eq("SI")
    open_known = verified & (~closed)
    comentarios = related_dfs.get("Comentarios_Incidencias", pd.DataFrame()).copy().fillna("")
    detalles = related_dfs.get("Detalles_Estados", pd.DataFrame()).copy().fillna("")

    comentarios_n3 = comentarios
    if not comentarios_n3.empty and {"Nivel", "Fecha_de_creacion"}.issubset(comentarios_n3.columns):
        comentarios_n3 = comentarios_n3[
            comentarios_n3["Nivel"].map(normalized).str.contains("NIVEL 3", na=False)
            & comentarios_n3["Fecha_de_creacion"].map(clean).ne("")
        ].copy()

    detalles_n3 = detalles
    if not detalles_n3.empty and {"Nivel", "Fecha_Estado"}.issubset(detalles_n3.columns):
        detalles_n3 = detalles_n3[
            detalles_n3["Nivel"].map(normalized).str.contains("NIVEL 3", na=False)
            & detalles_n3["Fecha_Estado"].map(clean).ne("")
        ].copy()

    comentario_campo = df[df["Comentario_Nivel3_Reporte"].map(clean).ne("")].copy()
    salesforce_id = first_nonempty(df, ["Salesforce_Id"])
    incidencia_id = first_nonempty(df, ["Incidencia_Id"])

    resumen_general = pd.DataFrame(
        [
            {"Concepto": "Pares procesados", "Cantidad": len(df)},
            {"Concepto": "Con Salesforce_Id", "Cantidad": int(salesforce_id.map(clean).ne("").sum())},
            {"Concepto": "Con Incidencia_Id", "Cantidad": int(incidencia_id.map(clean).ne("").sum())},
            {"Concepto": "Incidencias verificadas", "Cantidad": int(verified.sum())},
            {"Concepto": "Incidencias no verificadas", "Cantidad": int((~verified).sum())},
            {"Concepto": "Incidencias no cerradas verificadas", "Cantidad": int(open_known.sum())},
            {"Concepto": "Incidencias cerradas", "Cantidad": int(closed.sum())},
            {"Concepto": "No cerradas verificadas con cronograma", "Cantidad": int((open_known & has_cron).sum())},
            {"Concepto": "No cerradas verificadas sin cronograma", "Cantidad": int((open_known & (~has_cron)).sum())},
            {"Concepto": "Último cronograma activo", "Cantidad": int(df["Ultimo_Cronograma_Activo"].eq("SI").sum())},
            {"Concepto": "Último cronograma futuro", "Cantidad": int(df["Ultimo_Cronograma_Situacion"].eq("FUTURO").sum())},
            {"Concepto": "Último cronograma finalizado", "Cantidad": int(df["Ultimo_Cronograma_Situacion"].eq("FINALIZADO").sum())},
            {"Concepto": "Comentarios Nivel 3 con fecha", "Cantidad": len(comentarios_n3)},
            {"Concepto": "Detalles Estado Nivel 3 con fecha", "Cantidad": len(detalles_n3)},
            {"Concepto": "Comentario Nivel 3 en campo principal", "Cantidad": len(comentario_campo)},
        ]
    )

    estados = _summary_by(df, ["Estado_Incidencia_Reporte"])
    estados_n3 = _summary_by(df, ["Estado_Nivel3_Reporte", "Motivo_Nivel3_Reporte"])
    cron_fechas = _summary_by(df[df["Tiene_Cronograma"].eq("SI")], ["Cronograma_Inicio_Reporte", "Cronograma_Fin_Reporte"])

    return {
        "Resumen_General_Ext": resumen_general,
        "Resumen_Asignado": _summary_by(df, ["Asignado_Original"]),
        "Resumen_Zona": _summary_by(df, ["Zona_Reporte"]),
        "Resumen_Zona_Depto": _summary_by(df, ["Zona_Reporte", "Provincia_Reporte", "Departamento_Reporte"]),
        "Resumen_Departamento": _summary_by(df, ["Provincia_Reporte", "Departamento_Reporte"]),
        "Estados_Incidencia": estados,
        "Estados_Nivel3": estados_n3,
        "Cronogramas_Fecha": cron_fechas,
        "Resumen_Ultimo_Cronograma": _summary_by(df, ["Ultimo_Cronograma_Situacion", "Ultimo_Cronograma_Estado"]),
        "Ultimo_Crono_Activo": report_columns_first(df[df["Ultimo_Cronograma_Activo"].eq("SI")].copy()),
        "Ultimo_Crono_No_Activo": report_columns_first(df[df["Ultimo_Cronograma_Activo"].eq("NO")].copy()),
        "Abiertas_Con_Crono": report_columns_first(df[open_known & has_cron].copy()),
        "Abiertas_Sin_Crono": report_columns_first(df[open_known & (~has_cron)].copy()),
        "Cerradas_Aparte": report_columns_first(df[closed].copy()),
        "Sin_Incidencia_Id": report_columns_first(df[~verified].copy()),
        "Comentarios_Nivel3": comentarios_n3,
        "Detalles_Estados_N3": detalles_n3,
        "Comentario_N3_Campo": report_columns_first(comentario_campo),
    }


def create_input_template(path: str | Path, overwrite: bool = False) -> Path:
    path = Path(path)
    if path.exists() and not overwrite:
        return path
    plantilla = pd.DataFrame(
        [
            {
                "Numero_Predio": "",
                "Numero_Incidencia": "",
                "Asignado": "",
                "Zona": "",
                "Provincia": "",
                "Departamento": "",
                "Localidad_Ciudad": "",
                "Direccion": "",
                "GPS": "",
                "Notas": "",
            }
        ]
    )
    instrucciones = pd.DataFrame(
        [
            {"Campo": "Numero_Predio", "Uso": "Obligatorio. Numero de predio, por ejemplo 613668."},
            {"Campo": "Numero_Incidencia", "Uso": "Obligatorio. Incidencia, por ejemplo NI-000164106."},
            {"Campo": "Asignado", "Uso": "Opcional. Tecnico o responsable original de tu lista."},
            {"Campo": "Zona/Provincia/Departamento/Localidad_Ciudad", "Uso": "Opcional. El extractor lo preserva como Origen_*."},
            {"Campo": "Direccion/GPS/Notas", "Uso": "Opcional. Contexto original para reportes."},
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        plantilla.to_excel(writer, sheet_name="Entrada", index=False)
        instrucciones.to_excel(writer, sheet_name="Instrucciones", index=False)
        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            for col_idx, col_cells in enumerate(ws.columns, start=1):
                width = 12
                for cell in col_cells[:50]:
                    width = max(width, min(len(str(cell.value or "")) + 2, 55))
                ws.column_dimensions[get_column_letter(col_idx)].width = width
    return path


def format_workbook(writer: pd.ExcelWriter) -> None:
    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            width = 10
            for cell in col_cells[:150]:
                if cell.value is not None:
                    width = max(width, min(60, len(str(cell.value)) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
